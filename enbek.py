from pyPythonRPA.Robot import bySelector, keyboard, application, byImage
from pyPythonRPA import byDesk
import json
import os
from time import sleep
from os import listdir
from os.path import isfile, join
from xml.dom import minidom
# =======================================================================================================================================
import shutil
import glob
import time
import keyboard
from glob import glob
from pyPythonRPA.Robot import pythonRPA
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait, Select
from webdriver_manager.chrome import ChromeDriverManager
import datetime
import openpyxl
import json
import logging
import pathlib
# from Sources.winlog import WinLog
from winlog import WinLog
winlog = WinLog("HCSBKKZ_robot")

# with open((os.path.join(os.getcwd(), 'date_for_enbek.json'))) as colvir_json:
#     code = json.load(colvir_json)
# data_from_json = code["014169"][0]

# dogovor_number = data_from_json["dogovor_number"]
# date_dogovor = data_from_json["date_dogovor"]
# dolzhnost = data_from_json["dolzhnost"]
# mesto_raboty = data_from_json["mesto_raboty"]

# print(dogovor_number, date_dogovor, dolzhnost, mesto_raboty)


def extract_dolzhnosts_from_excel():
    import openpyxl
    import json
    path = r"C:\Users\robot.drp\Desktop\Rasul\HCSBKKZ_robot\Tools\sample.xlsx"

    wb_obj = openpyxl.load_workbook(path)

    # Read the active sheet:
    sheet = wb_obj.active

    winlog.info(f"sheet {sheet}")
    dict_shtat_enbek = {}
    dict_for_json = {}

    for row in sheet.iter_rows():

        podrazdelenye = row[1].value
        if podrazdelenye:
            podrazdelenye = podrazdelenye.lower()

        shtat_dolzh = row[2].value
        if shtat_dolzh:
            shtat_dolzh = shtat_dolzh.lower()

        enbek_dolzh = row[4].value
        if not enbek_dolzh:
            enbek_dolzh = "Разработчик программного обеспечения"
        #
        dict_shtat_enbek[shtat_dolzh] = enbek_dolzh
        if podrazdelenye in dict_for_json:
            dict_for_json[podrazdelenye].append(dict_shtat_enbek)
            dict_shtat_enbek = {}
        else:
            dict_for_json[podrazdelenye] = [dict_shtat_enbek]
            dict_shtat_enbek = {}

        # for i, cell in enumerate(row):
        #     if i == 1 or i == 2 or i == 4:
        #         if cell.value == None:
        #             print(cell.value, end="+" * 10)
        #         else:
        #             print(cell.value.lower(), end="+" * 10)
        # print()
        # print()

    with open("dolzhonst_with_filter.json", "w", encoding="utf-8") as file:
        json.dump(dict_for_json, file, indent=4, ensure_ascii=False)


class Enbek:

    def __init__(self):
        self.robot_path = os.getcwd()
        self.driver = None
        self.downloads_path = os.path.join(os.environ['USERPROFILE'], "Downloads\\").replace("/", "\\")
        self.temp_path = self.robot_path + "Temp\\Enbek"
        self.path = {"downloads": {"title": "Downloads", "path": self.downloads_path, "dir": self.downloads_path[:-1]},
                     "temp": {"title": "Temp", "path": self.temp_path, "dir": self.temp_path[:-1]},
                     "enbek_files": {"title": "Enbek_files", "path": self.temp_path + "Enbek_files\\",
                                     "dir": self.temp_path + "Enbek_files"}}
        self.url = {
            "login": "https://www.enbek.kz/docs/ru/user",
            "list": "https://www.enbek.kz/ru/cabinet/dogovor/list/good",
            "add": "https://www.enbek.kz/ru/cabinet/dogovor/add",
        }
        # Data containers
        self.anchor = None
        winlog.info("\t")

    def _sel_init(self):
        """Запуск драйвера и логин в enbek. Объект драйвера создается именно здесь"""
        # Driver init
        try:
            self.driver = webdriver.Chrome(ChromeDriverManager().install())
        except Exception as e:
            winlog.warning(f"{datetime.datetime.today()} {'-' * 10} _sel_init {e}")
            self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.switch_to.window(self.driver.current_window_handle)
        driver = self.driver
        driver.get(self.url["list"])
        login = "//input[@placeholder='Логин или E-mail']"
        WebDriverWait(driver, 60).until(ec.presence_of_element_located((By.XPATH, login)))
        driver.find_element(By.XPATH, login).send_keys("madieva.an@hcsbk.kz")
        passw = "//input[@placeholder='Пароль']"
        driver.find_element(By.XPATH, passw).send_keys("Ghjdthrf_2021")
        driver.find_element(By.XPATH, passw).send_keys(Keys.RETURN)
        driver.get(self.url["list"])
        return self.driver

    def create_dogovor(self, data):
        if not self.driver:
            self._sel_init()

        iin_exist = self._check_dogovor_create(data)
        if not iin_exist:
            self._fill_dog(data)
        time.sleep(3)
        # self.driver.quit()

    def _check_dogovor_create(self, data):
        driver = self.driver
        driver.get(self.url["list"])
        if self._sel_wait_el(By.XPATH, "//a[text()[contains(., 'Добавить')]]"):
            self.anchor = driver.find_element(By.XPATH, '//strong[text()="Договоры"]')
            root = driver.find_element(By.CSS_SELECTOR, ".content")

            iin_iin = data["IIN"]
            iin_exist = self._search_iin_create(root, iin_iin)
            winlog.info(f"{iin_exist}")
        else:
            raise ValueError("Время ожидания истекло: https://www.enbek.kz/ru/cabinet/dogovor/list")
        winlog.info("self._check_dogovor > done")
        return iin_exist

    def _search_iin_create(self, root, iin):
        self.anchor.click()

        input_iin = root.find_element(By.XPATH, ".//input[@name='iin']")
        input_iin.send_keys(iin)
        button_iin = root.find_element(By.XPATH,
                                       './/button[@type="submit" and text()="Найти"]')
        button_iin.click()

        if self._sel_wait_el(By.XPATH, '//strong[text()="Пусто..." ]', 3):
            return False
        else:
            return True

    def _sel_wait_el(self, by, selector, sec=60, appear=True):
        """Ожидание элемента появление или изчезновение подается через bool 'appear'"""
        driver = self.driver
        time.sleep(0.3)
        try:
            if appear:
                WebDriverWait(driver, sec).until(ec.presence_of_element_located((by, selector)))
            else:
                WebDriverWait(driver, sec).until_not(ec.presence_of_element_located((by, selector)))
            return True
        except:
            return False
        finally:
            time.sleep(0.2)

    def _check_iin(self, root, counter=1):
        while counter:
            self._sel_wait_el(By.XPATH,
                              ".//input[@name='IIN']/parent::div/parent::div/div[@class='mdb-ehr-progress' and @style='display: block;']",
                              5)
            self._sel_wait_el(By.XPATH,
                              ".//input[@name='IIN']/parent::div/parent::div/div[@class='mdb-ehr-progress' and @style='display: none;']",
                              5)
            time.sleep(1)
            if counter > 0:
                input_fam = root.find_element(By.XPATH, ".//input[@name='FAM']")
                if len(input_fam.get_attribute("value")) > 1:
                    print(input_fam.get_attribute("value"))
                    return True
            else:
                counter -= 1

        return False

    def _fill_iin(self, root, iin):
        self.anchor.click()

        input_iin = root.find_element(By.XPATH, ".//input[@name='IIN']")
        input_iin.send_keys(iin)
        button_iin = root.find_element(By.XPATH, ".//input[@name='IIN']/parent::div/span/button")
        button_iin.click()
        flag = self._check_iin(root, 2)
        if not flag:
            raise ValueError("Время ожидания истекло: не найден по ИИН")
        winlog.info("self._fill_iin > done")

    def _fill_string(self, root, num, shtatka_full):
        self.anchor.click()

        input_num = root.find_element(By.XPATH, ".//input[@name='numDogovor']")
        input_num.send_keys(num)

        # redefinition

        input_dol = root.find_element(By.XPATH, ".//input[@name='shtatDolj']")
        input_dol.send_keys(shtatka_full)

        winlog.info("self._fill_string > done")

    def _fill_select(self, root, srok, vid):
        self.anchor.click()

        select_srok = Select(root.find_element(By.XPATH, ".//select[@name='dContractCate']"))
        select_srok.select_by_visible_text(srok)

        select_vid = Select(root.find_element(By.XPATH, ".//select[@name='partTime']"))
        select_vid.select_by_visible_text(vid)

        winlog.info("self._fill_select > done")

    def _fill_rezhim(self, root, rezhim, stavka=False):
        self.anchor.click()

        select_rezhim = Select(root.find_element(By.XPATH, ".//select[@name='workingHours']"))
        select_rezhim.select_by_visible_text(rezhim)

        if stavka:
            self._sel_wait_el(By.XPATH, ".//input[@name='tariffRate']")
            input_stavka = root.find_element(By.XPATH, ".//input[@name='tariffRate']")
            input_stavka.send_keys(stavka)

        winlog.info("self._fill_rejim > done")

    def _fill_date(self, root, dogovor, nachalo, konec=""):
        self.anchor.click()

        input_dogovor = root.find_element(By.XPATH, ".//input[@name='dateZakDogovor']")
        input_dogovor.click()
        input_dogovor.send_keys(dogovor)
        self.anchor.click()

        input_nachalo = root.find_element(By.XPATH, ".//input[@name='dateBegDogovor']")
        input_nachalo.click()
        input_nachalo.send_keys(nachalo)
        self.anchor.click()

        if len(konec):
            input_konec = root.find_element(By.XPATH, ".//input[@name='dateEndDogovor']")
            input_konec.click()
            input_konec.send_keys(konec)
            self.anchor.click()

        winlog.info("self._fill_date > done")

    def _fill_dol(self, root, dol):
        driver = self.driver
        self.anchor.click()

        span_dogovor = root.find_element(By.XPATH,
                                         ".//label[text()='Должность ']/parent::div//span[@class='selection']")
        span_dogovor.click()

        root_dogovor = driver.find_element(By.XPATH,
                                           "//span[@class='select2-container select2-container--default select2-container--open']")

        input_dogovor = root_dogovor.find_element(By.XPATH, ".//input[@class='select2-search__field']")

        input_dogovor.send_keys(dol)

        if self._sel_wait_el(By.XPATH,
                             ".//li[@class='select2-results__option select2-results__option--highlighted' and text()='" + dol + "']",
                             sec=5):
            li_dogovor = driver.find_element(By.XPATH,
                                             ".//li[@class='select2-results__option select2-results__option--highlighted' and text()='" + dol + "']")
        elif self._sel_wait_el(By.XPATH, ".//li[@class='select2-results__option' and text()='" + dol + "']", sec=5):
            li_dogovor = driver.find_element(By.XPATH,
                                             ".//li[@class='select2-results__option' and text()='" + dol + "']")
        else:
            raise ValueError("Время ожидания истекло: Должность не найдена")
        li_dogovor.click()
        if not self._sel_wait_el(By.CSS_SELECTOR, "span.select2-container--open input.select2-search__field",
                                 appear=False):
            raise ValueError("Время ожидания истекло: Должность не выбрана")

        winlog.info("self._fill_dol > done")

    def _fill_adres(self, root, obl, center, adres, f):
        driver = self.driver
        self.anchor.click()

        button_obl = root.find_element(By.XPATH, ".//Button[text()='Выбрать']")
        button_obl.click()

        self._sel_wait_el(By.XPATH, ".//div[@class='modal-content' and //h4[text()='Справочник регионов']]")
        root_adres = driver.find_element(By.XPATH,
                                         ".//div[@class='modal-content' and //h4[text()='Справочник регионов']]")

        li_obl = root_adres.find_element(By.XPATH, ".//li[span[text()='" + obl + "']]")
        li_obl.click()

        time.sleep(1)
        self._sel_wait_el(By.XPATH, ".//li[span[text()='" + center + "']]")
        li_center = root_adres.find_element(By.XPATH, ".//li[span[text()='" + center + "']]")
        li_center.click()

        button_adres = root_adres.find_element(By.XPATH, ".//button[text()='Выбор']")
        button_adres.click()
        if not self._sel_wait_el(By.XPATH, ".//div[@class='modal-content' and //h4[text()='Справочник регионов']]",
                                 appear=False):
            raise ValueError("Время ожидания истекло: Адрес не выбран")

        input_adres = root.find_element(By.XPATH, ".//input[@name='workPlace']")
        input_adres.send_keys(adres)

        input_mesto_do_work = driver.find_element(By.XPATH, '//input[@name="workPlace"]')
        input_mesto_do_work.send_keys(f)

        # if f:
        #     nas_punkt = 'г.'+center
        #     span_nas_punkt = root.find_element(By.XPATH,
        #                                        ".//label[text()='Населённый пункт ']/parent::div//span[@class='selection']")
        #     span_nas_punkt.click()
        #
        #     root_nas = driver.find_element(By.XPATH,
        #                                    "//span[@class='select2-container select2-container--default select2-container--open']")
        #     input_nas_punkt = root_nas.find_element(By.XPATH, ".//input[@class='select2-search__field']")
        #     # input_nas_punkt.send_keys(nas_punkt)
        #     input_nas_punkt.send_keys(f)
        #
        #     if self._sel_wait_el(By.XPATH,
        #                          ".//li[@class='select2-results__option select2-results__option--highlighted' and text()='" + f + "']",
        #                          sec=5):
        #         li_nas = driver.find_element(By.XPATH,
        #                                      ".//li[@class='select2-results__option select2-results__option--highlighted' and text()='" + f + "']")
        #     elif self._sel_wait_el(By.XPATH, ".//li[@class='select2-results__option' and text()='" + f + "']",
        #                            sec=5):
        #         li_nas = driver.find_element(By.XPATH,
        #                                      ".//li[@class='select2-results__option' and text()='" + f + "']")
            # else:
            #     raise ValueError("Время ожидания истекло: Должность не найдена")
            # li_nas.click()
            # if not self._sel_wait_el(By.CSS_SELECTOR, "span.select2-container--open input.select2-search__field",
            #                          appear=False):
            #     raise ValueError("Время ожидания истекло: Должность не выбрана")

        winlog.info("self._fill_adres > done")

    def _fill_dog(self, data_from_json):
        driver = self.driver
        driver.get(self.url["add"])
        if self._sel_wait_el(By.XPATH, "//h3[text()='Добавление договора']"):
            self.anchor = driver.find_element(By.XPATH, "//h3[text()='Добавление договора']")
            root = driver.find_element(By.CSS_SELECTOR, ".content")

            iin_iin = data_from_json["IIN"]
            self._fill_iin(root, iin_iin)

            string_num = data_from_json["dogovor_number"]
            string_dol = data_from_json["dolzhnost"].strip()  # Штатная доджность
            shtatka_full = data_from_json["department"].strip() + " " + data_from_json["podrazdelenye"].strip() + " " + data_from_json["dolzhnost"]
            self._fill_string(root, string_num, shtatka_full)

            select_srok = data_from_json["srok_dogovor"]
            select_vid = data_from_json["work_type"]
            self._fill_select(root, select_srok, select_vid)

            # rezhim = data_from_json["part_time"].split(", ")
            # rezhim_rezhim = rezhim[0]
            # rezhim_stavka = rezhim[1] if len(rezhim) > 1 else False
            # self._fill_rezhim(root, rezhim_rezhim, rezhim_stavka)

            date_dogovor = data_from_json["date_dogovor"]
            date_nachalo = data_from_json["date_dogovor"]
            date_konec = data_from_json["date_dogovor"][:-1] + str(int(data_from_json["date_dogovor"][-1]) + 1) if (
                    "на определенный срок не менее одного года" in select_srok or "на время выполнения сезонной работы" in select_srok) else ""
            self._fill_date(root, date_dogovor, date_nachalo, date_konec)


            with open("dolzhonst_with_filter.json",  encoding="utf-8") as file:
                dict_ = json.load(file)
            winlog.info(dict_)
            # colv_podr = "Управление учета финансовых инструментов"
            # colv_dolzh = "Заместитель начальника"
            colv_podr = data_from_json["podrazdelenye"]
            colv_dolzh = data_from_json["dolzhnost"]
            winlog.info(f"colv_dolzh {colv_dolzh}", )
            winlog.info(f"colv_podr {colv_podr}")
            # enbek_dolzh_ = "Разработчик программного обеспечения"
            # enbek_dolzh_ = "Разработчик программного обеспечения"
            enbek_dolzh_ = ""
            for podr in dict_:
                if colv_podr.lower() == podr:
                    for i, dict_sht_enb in enumerate(dict_[
                                                         podr]):  # каждый словарь у которого ключ: штатная должность, значение: должность на енбеке
                        if colv_dolzh.lower() in dict_sht_enb:
                            enbek_dolzh_ = dict_[podr][i][colv_dolzh.lower()]

            # print(enbek_dolzh_)
            # dol_dol = data_from_json["dolzhnost"] # ------------------------------------------
            dol_dol = enbek_dolzh_
            self._fill_dol(root, dol_dol)

            select_country = Select(root.find_element(By.XPATH, ".//select[@name='workPlaceCountry']"))
            select_country.select_by_visible_text('Казахстан')

            # adres_obl = data_from_json["Место выполнения работы"]["obl"]
            # adres_center = data_from_json["Место выполнения работы"]["city"]
            # adres_adres = data_from_json["Место выполнения работы"]["street"]
            # print(adres_obl, adres_center, adres_adres)
            adres_main = data_from_json["mesto_raboty"]
            f = False
            # if adres_obl == 'УСТЬ-КАМЕНОГОРСК':
            #     adres_obl = 'В-КАЗАХСТАНСКАЯ ОБЛАСТЬ'
            #     f = True
            # self._fill_adres(root, adres_obl, adres_center, adres_adres, f)
            # TODO change to back and ADD more elifs for each city
            # dfa = data_for_adres
            if data_from_json["mesto_raboty"] == "ЦА":
                dfa = self._map_location()['АО "Жилищный строительный сберегательный банк "Отбасы банк"']
            else:
                dfa = self._map_location()['Филиал АО "Жилищный строительный сберегательный банк "Отбасы банк" в городе Алматы']
            self._fill_adres(root, dfa[0], dfa[1], dfa[2], dfa[3])
            # select_military = Select(root.find_element(By.XPATH, ".//select[@name='army']"))
            # select_military.select_by_visible_text(data["Военная обязанность"])
            # input("before save")
            # try:
            #     submit = root.find_element(By.XPATH, ".//input[@value='Сохранить']")
            #     submit.click()
            # except Exception as e:
            driver.execute_script('document.getElementsByClassName("btn btn-primary submit-form")[0].click()')
            # TODO add function for applye dogovor
            self.apply_and_send_dogovor()
            self.ecp_priem()
            # input("stop before one apply")
            sleep(1)
            # go to dogovory for next priem
            # driver.find_element(By.XPATH, '//a[@href="/ru/cabinet/dogovor/list/all"]').click()
            driver.get("https://www.enbek.kz/ru/cabinet/dogovor/list/all")
        else:
            raise ValueError("Время ожидания истекло: https://www.enbek.kz/ru/cabinet/dogovor/add")
        winlog.info("self._fill_dog > done")

    def apply_and_send_dogovor(self):
        driver = self.driver
        if self._sel_wait_el(By.XPATH, '//button[text()="Подписать договор и отправить"]'):
            driver.find_element(By.XPATH, '//button[text()="Подписать договор и отправить"]').click()
            self._sel_wait_el(By.XPATH, '//button[text()="OK"]')
            driver.find_element(By.XPATH, '//button[text()="OK"]').click()

    def ecp_priem(self):
        file_name_input = bySelector([{"title": "Электронная биржа труда | ID ДОГОВОРА (E-HR): - Google Chrome", "backend": "uia"},
                                      {"depth_start": 3, "depth_end": 3, "title": "File name:",
                                       "control_type": "Edit"}])
        file_name_input.wait_appear()
        file_name_input.set_focus()
        sleep(0.5)
        file_name_input.set_edit_text(
            r"C:\Users\robot.drp\Desktop\Rasul\HCSBKKZ_robot\Tools\ecp\GOSTKNCA_42e6bd1a0979ff3747e9a35207bbcfbc4afee6f5.p12")
        sleep(0.5)

        open_button = bySelector([{"title": "Электронная биржа труда | ID ДОГОВОРА (E-HR): - Google Chrome", "backend": "uia"},
                                  {"depth_start": 2, "depth_end": 2, "title": "Open", "control_type": "Button"}])
        open_button.click()

        driver = self.driver
        self._sel_wait_el(By.XPATH, '(//input[@name="ecpPassword"])[1]')
        driver.find_element(By.XPATH, '(//input[@name="ecpPassword"])[3]').send_keys("test_password")
        driver.find_element(By.XPATH, '(//button[contains(text(), "OK")])[3]').click()
        # input("def ecp_priem")

    def _map_location(self):
        location_mapping = {
            'Филиал АО "Жилищный строительный сберегательный банк "Отбасы банк" в городе Семей':
                ['В-КАЗАХСТАНСКАЯ ОБЛАСТЬ', 'Семей', 'г.Семей', 'ул. Чайжунусова, 152А'],
            'Костанайский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"':
                ['КОСТАНАЙСКАЯ ОБЛАСТЬ', 'Костанай', 'г.Костанай', 'пр. аль-Фараби, 67'],
            'Актюбинский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"':
                ['АКТЮБИНСКАЯ ОБЛАСТЬ', 'Актобе', 'г.Актобе', 'пр. А. Молдагуловой, 46Б'],
            'Филиал АО "Жилищный строительный сберегательный банк "Отбасы банк" в городе Шымкент':
                ['ШЫМКЕНТ', 'Шымкент', 'г.Шымкент', 'ул. К. Рыскулбекова, 3Г'],
            'Туркестанский ОФ АО "Жилищный строительный сберегательный банк "Отбасы Банк"':
                ['ТУРКЕСТАНСКАЯ ОБЛАСТЬ', 'Туркестан', 'г.Туркестан', 'пр. Б.Саттарханова 146/22'],
            'Жамбылский ОФ АО "Жилищный строительный сберегательный банк "Отбасы Банк"':
                ['ЖАМБЫЛСКАЯ ОБЛАСТЬ', 'Тараз', 'г.Тараз', 'пр. Жамбыла, 9Б'],
            'Северо-Казахстанский ОФ АО "Жилищный строительный сберегательный банк "Отбасы Банк"':
                ['С-КАЗАХСТАНСКАЯ ОБЛАСТЬ', 'Петропавловск', 'г.Петропавловск', 'ул. Конституции Казахстана, 28'],
            'Кызылординский ОФ АО "Жилищный строительный сберегательный банк "Отбасы Банк"':
                ['КЫЗЫЛОРДИНСКАЯ ОБЛАСТЬ', 'Кызылорда', 'г.Кызылорда', 'ул. Байтурсынова, 122'],
            'Мангистауский ОФ АО «Жилищный строительный сберегательный банк «Отбасы Банк»':
                ['МАНГИСТАУСКАЯ ОБЛАСТЬ', 'Актау', 'г.Актау', 'мкр. 11, 60'],
            'Атырауский ОФ АО "Жилищный строительный сберегательный банк "Отбасы Банк"':
                ['АТЫРАУСКАЯ ОБЛАСТЬ', 'Атырау', 'г.Атырау', 'ул. Абая, 15А/1'],
            'Карагандинский ОФ АО "Жилищный строительный сберегательный банк "Отбасы Банк"':
                ['КАРАГАНДИНСКАЯ ОБЛАСТЬ', 'Караганда', 'г.Караганда',
                 'район имени Казыбек би, пр. Бухар Жырау, стр. 57/1'],
            'Павлодарский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"':
                ['ПАВЛОДАРСКАЯ ОБЛАСТЬ', 'Павлодар', 'г.Павлодар', 'ул. Абая, 75'],
            'Западно-Казахстанский ОФ АО "Жилищный строительный сберегательный банк "Отбасы Банк"':
                ['З-КАЗАХСТАНСКАЯ ОБЛАСТЬ', 'Уральск', 'г.Уральск', 'ул. С.Датова, 28'],
            'Акмолинский ОФ АО «Жилищный строительный сберегательный банк «Отбасы банк»':
                ['АКМОЛИНСКАЯ ОБЛАСТЬ', 'Кокшетау', 'г.Кокшетау', 'ул. 8 марта, 51'],
            'Восточно-Казахстанский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"':
                ['В-КАЗАХСТАНСКАЯ ОБЛАСТЬ', 'Усть-Каменогорск', 'г.Усть-Каменогорск', 'ул. Головкова, 25/1'],
            'Центральный Ф АО "Жилищный строительный сберегательный банк "Отбасы банк"':
                ['НУР-СУЛТАН', 'район Алматы', 'район Алматы', 'пр. Р.Қошкарбаева, д. 26, н.п 16'],
            'Алматинский ОФ АО "Жилищный строительный сберегательный банк "Отбасы банк"':
                ['АЛМАТИНСКАЯ ОБЛАСТЬ', 'Талдыкорган', 'г.Талдыкорган', 'ул. М.Толебаева, 86'],
            'Филиал АО "Жилищный строительный сберегательный банк "Отбасы банк" в городе Алматы':
                ['АЛМАТЫ', 'Алмалинский район', 'Алмалинский район', 'пр. Сейфуллина, 498'],
            'АО "Жилищный строительный сберегательный банк "Отбасы банк"':
                ['АЛМАТЫ', 'Алмалинский район', 'Алмалинский район', 'пр. Абылайхана 91']
        }
        return location_mapping

# ==================================================================================================================================================================


class EnbekPerevod:
    def __init__(self):
        self.robot_path = os.getcwd()
        self.driver = None
        self.downloads_path = os.path.join(os.environ['USERPROFILE'], "Downloads\\").replace("/", "\\")
        self.temp_path = self.robot_path + "Temp\\Enbek"
        self.path = {"downloads": {"title": "Downloads", "path": self.downloads_path, "dir": self.downloads_path[:-1]},
                     "temp": {"title": "Temp", "path": self.temp_path, "dir": self.temp_path[:-1]},
                     "enbek_files": {"title": "Enbek_files", "path": self.temp_path + "Enbek_files\\",
                                     "dir": self.temp_path + "Enbek_files"}}
        self.url = {
            "login": "https://www.enbek.kz/docs/ru/user",
            "list": "https://www.enbek.kz/ru/cabinet/dogovor/list/good",
            "add": "https://www.enbek.kz/ru/cabinet/dogovor/add",
        }
        # Data containers
        self.anchor = None
        winlog.info("\t")

    def ecp_priem(self, count_dopka: int):
        file_name_input = bySelector([{"title": "Электронная биржа труда | ID ДОГОВОРА (E-HR): - Google Chrome", "backend": "uia"},
                                      {"depth_start": 3, "depth_end": 3, "title": "File name:",
                                       "control_type": "Edit"}])
        file_name_input.wait_appear()
        file_name_input.set_focus()
        sleep(0.5)
        file_name_input.set_edit_text(
            r"C:\Users\robot.drp\Desktop\Rasul\HCSBKKZ_robot\Tools\ecp\GOSTKNCA_42e6bd1a0979ff3747e9a35207bbcfbc4afee6f5.p12")
        sleep(0.5)

        open_button = bySelector([{"title": "Электронная биржа труда | ID ДОГОВОРА (E-HR): - Google Chrome", "backend": "uia"},
                                  {"depth_start": 2, "depth_end": 2, "title": "Open", "control_type": "Button"}])
        open_button.click()

        driver = self.driver
        self._sel_wait_el(By.XPATH, '(//input[@name="ecpPassword"])[1]')
        # такие странные операции с количеством допок, потому что там саит загружает для всех допок эти инпуты и кнопки, даже если они не были выбраны
        winlog.info(f"before find element count_dopka {count_dopka}")

        if count_dopka == 0:
            winlog.info("launch part apply ecp for first dopka")
            driver.find_element(By.XPATH, '(//input[@name="ecpPassword"])[3]').send_keys("test_password")
            driver.find_element(By.XPATH, '(//button[contains(text(), "OK")])[3]').click()
        else:
            winlog.info("launch part apply ecp for non first dopka")
            driver.find_element(By.XPATH, f'(//input[@name="ecpPassword"])[{count_dopka * 2 + 1}]').send_keys(
                "test_password")
            driver.find_element(By.XPATH, f'(//button[contains(text(), "OK")])[{count_dopka * 2 + 1}]').click()

        driver.get("https://enbek.kz/ru/cabinet/dogovor/list/all")
        # a = input("a = ecp_priem")

    @staticmethod
    def extract_data_from_excel_for_perevod(): # and write to json
        path = r"C:\Users\robot.drp\Desktop\Rasul\HCSBKKZ_robot\Tools\sample.xlsx"

        wb_obj = openpyxl.load_workbook(path)

        # Read the active sheet:
        sheet = wb_obj.active

        winlog.info(sheet)
        dict_shtat_enbek = {}
        dict_for_json = {}

        for row in sheet.iter_rows():

            department = row[0].value
            if department:
                department = department.lower().strip()

            shtat_dolzh = row[2].value
            if shtat_dolzh:
                shtat_dolzh = shtat_dolzh.lower().strip()

            enbek_dolzh = row[4].value
            if not enbek_dolzh:
                enbek_dolzh = "Разработчик программного обеспечения"
            #
            dict_shtat_enbek[shtat_dolzh] = enbek_dolzh.strip()
            if department in dict_for_json:
                dict_for_json[department].append(dict_shtat_enbek)
                dict_shtat_enbek = {}
            else:
                dict_for_json[department] = [dict_shtat_enbek]
                dict_shtat_enbek = {}

            # for i, cell in enumerate(row):
            #     if i == 1 or i == 2 or i == 4:
            #         if cell.value == None:
            #             print(cell.value, end="+" * 10)
            #         else:
            #             print(cell.value.lower(), end="+" * 10)
            # print()
            # print()

        with open("dolzhonst_with_filter_for_perevod.json", "w", encoding="utf-8") as file:
            json.dump(dict_for_json, file, ensure_ascii=False, indent=4)

    def _sel_init(self):
        """Запуск драйвера и логин в enbek. Объект драйвера создается именно здесь"""
        # Driver init
        try:
            self.driver = webdriver.Chrome(ChromeDriverManager().install())
        except Exception as e:
            winlog.warning(f"{datetime.datetime.today()} {'-' * 10} _sel_init_2 {e}")
            self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.switch_to.window(self.driver.current_window_handle)
        driver = self.driver
        driver.get(self.url["list"])
        login = "//input[@placeholder='Логин или E-mail']"
        WebDriverWait(driver, 60).until(ec.presence_of_element_located((By.XPATH, login)))
        driver.find_element(By.XPATH, login).send_keys("madieva.an@hcsbk.kz")
        passw = "//input[@placeholder='Пароль']"
        driver.find_element(By.XPATH, passw).send_keys("Ghjdthrf_2021")
        driver.find_element(By.XPATH, passw).send_keys(Keys.RETURN)
        driver.get(self.url["list"])
        return self.driver

    def create_dop_sogl(self, data):
        if not self.driver:
            self._sel_init()

        link, count_dopka = self._check_dop_sogl(data)

        status_dolzh_found, department, dolzhnost = self._data_append(data, link, count_dopka)
        return status_dolzh_found, department, dolzhnost

    def _check_dop_sogl(self, data):
        driver = self.driver
        driver.get(self.url["list"])

        if self._sel_wait_el(By.XPATH, "//a[text()[contains(., 'Добавить')]]"):
            self.anchor = driver.find_element(By.XPATH, '//strong[text()="Договоры"]')
            root = driver.find_element(By.CSS_SELECTOR, ".content")

            iin_iin = data[-1]
            link, count_dop_sogl = self._search_iin_create1(root, iin_iin, driver)
            winlog.info("self._check_dogovor > done")
            return link, count_dop_sogl
        else:
            winlog.info("self._check_dogovor > done")
            raise ValueError("Время ожидания истекло: https://www.enbek.kz/ru/cabinet/dogovor/list")

    def _data_append(self, data, link, count_dopka):
        root = self.driver
        root.get(link)
        no_dop_sogl = root.find_element_by_xpath('//input[@name="numDogovor"]')

        no_dop_sogl.send_keys(f"{len(count_dopka) + 1}")

        # select_rezhim = Select(root.find_element(By.XPATH, "//select[@name='workingHours']"))
        # select_rezhim.select_by_visible_text(data["Режим рабочего времени"])

        date_nachalo_dop_sogl = root.find_element_by_xpath("//input[@name='dateBegDogovor']")
        date_nachalo_dop_sogl.click()
        date_list = data[3].split(".")
        date_dopka = date_list[0] + "." + date_list[1] + "." + str(datetime.datetime.today())[:2] + date_list[2]

        date_nachalo_dop_sogl.send_keys(date_dopka)

        date_zakl_dop_sogl = root.find_element_by_xpath("//input[@name='dateZakDogovor']")
        date_zakl_dop_sogl.click()
        date_zakl_dop_sogl.send_keys(date_dopka)

        # data['Срок договора']
        select_srok = Select(root.find_element(By.XPATH, "//select[@name='srokdop']"))
        print("seem data before srok", data)
        if data[-2] == "Постоянная":
            select_srok.select_by_visible_text("на неопределенный срок")
        else:
            select_srok.select_by_visible_text("на время замещения временно отсутствующего работника")

        # dol = data[2]
        # dol = "IT-дизайнер"
        # ======================
        with open((os.path.join(os.getcwd(), 'data_for_perevod.json')), encoding="utf-8") as colvir_json:
            code = json.load(colvir_json)

        # print(data)  # ['Департамент безопасности', 'Сектор безопасности по Южному региону', 'Специалист 1 Категории', '07.10.21', 'Департамент безопасности', 'Сектор безопасности по Южному региону', 'Заведующий Сектором', '650223301491']
        with open("dolzhonst_with_filter_for_perevod.json", encoding="utf-8") as file:
            dict_ = json.load(file)

        # print(dict_)

        # colv_podr = "Управление учета финансовых инструментов"
        # colv_dolzh = "Заместитель начальника"
        excel_department = data[-5]
        winlog.info(f"excel_department {excel_department}")
        colv_dolzh = data[-3]
        winlog.info(f"colv_dolzh {colv_dolzh}")

        # enbek_dolzh_ = "Разработчик программного обеспечения"
        # enbek_dolzh_ = "Разработчик программного обеспечения"
        enbek_dolzh_ = ""
        flag_main_cycle = 0
        for department in dict_:
            if excel_department.lower() == department:
                winlog.info(f"department for perevod is found {department}")
                for i, dict_sht_enb in enumerate(dict_[department]):  # каждый словарь у которого ключ: штатная должность, значение: должность на енбеке
                    winlog.info(f"dict sht : dolzh {dict_sht_enb}")
                    if colv_dolzh.lower() in dict_sht_enb:
                        winlog.info("Dolzhnost found")
                        enbek_dolzh_ = dict_[department][i][colv_dolzh.lower()]
                        flag_main_cycle = 1
                        break
            if flag_main_cycle == 1:
                break

        dol_dol = enbek_dolzh_

        # =================
        span_dogovor = root.find_element(By.XPATH,"//span[@class='selection']")
        span_dogovor = span_dogovor.find_element(By.XPATH, 'span')
        span_dogovor.click()
        root_dogovor = root.find_element(By.XPATH,"//span[@class='select2-container select2-container--default select2-container--open']")
        input_dogovor = root_dogovor.find_element(By.XPATH, ".//input[@class='select2-search__field']")
        input_dogovor.send_keys(dol_dol)
        # input("fix input dolzh")
        li_dogovor = ""
        print("dol dol ", dol_dol)
        if self._sel_wait_el(By.XPATH,
                             "//li[@class='select2-results__option select2-results__option--highlighted' and text()='" + dol_dol + "']",
                             sec=5):
            li_dogovor = root.find_element(By.XPATH,
                                             "//li[@class='select2-results__option select2-results__option--highlighted' and text()='" + dol_dol + "']")
        elif self._sel_wait_el(By.XPATH, "//li[@class='select2-results__option' and text()='" + dol_dol + "']", sec=5):
            li_dogovor = root.find_element(By.XPATH,
                                             "//li[@class='select2-results__option' and text()='" + dol_dol + "']")
        else:
            # raise ValueError("Время ожидания истекло: Должность не найдена")
            winlog.info(f"Время ожидания истекло: Нет данных по: Департамент: {excel_department} Штатная должность: {colv_dolzh}")
            return 0, excel_department, colv_dolzh
        li_dogovor.click()
        if not self._sel_wait_el(By.CSS_SELECTOR, "span.select2-container--open input.select2-search__field",
                                 appear=False):
            raise ValueError("Время ожидания истекло: Должность не выбрана")

        # data['Штатная должность']
        input_dol = root.find_element(By.XPATH, "//input[@name='shtatDolj']")
        # input_dol.send_keys(data['Штатная должность'])
        input_dol.send_keys((data[-3] + " " + data[-4] + " " + data[-5]))
        # data['Место выполнения работы']
        # obl = data["Место выполнения работы"]["obl"]
        # center = data["Место выполнения работы"]["city"]
        # if obl == 'УСТЬ-КАМЕНОГОРСК':
        #     obl = 'В-КАЗАХСТАНСКАЯ ОБЛАСТЬ'
        # adres = data["Место выполнения работы"]["street"]
        # select_country = Select(root.find_element(By.XPATH, "//select[@name='workPlaceCountry']"))
        # select_country.select_by_visible_text('Казахстан')
        # button_obl = root.find_element(By.XPATH, "//Button[text()='Выбрать']")
        # button_obl.click()

        # self._sel_wait_el(By.XPATH, "//div[@class='modal-content' and //h4[text()='Справочник регионов']]")
        # root_adres = root.find_element(By.XPATH,
        #                                  ".//div[@class='modal-content' and //h4[text()='Справочник регионов']]")
        # li_obl = root_adres.find_element(By.XPATH, ".//li[span[text()='" + obl + "']]")
        # li_obl.click()
        # self._sel_wait_el(By.XPATH, "//li[span[text()='" + center + "']]")
        # time.sleep(1)
        # li_center = root_adres.find_element(By.XPATH, "//li[span[text()='" + center + "']]")
        # time.sleep(1)
        # li_center.click()
        #
        # button_adres = root_adres.find_element(By.XPATH, ".//button[text()='Выбор']")
        # button_adres.click()
        # if not self._sel_wait_el(By.XPATH, "//div[@class='modal-content' and //h4[text()='Справочник регионов']]",
        #                          appear=False):
        #     raise ValueError("Время ожидания истекло: Адрес не выбран")
        #
        # input_adres = root.find_element(By.XPATH, "//input[@name='workPlace']")
        # input_adres.send_keys(adres)
        #
        # span_nas = root.find_element(By.XPATH, ".//label[text()='Населённый пункт ']/parent::div//span[@class='selection']")
        # span_nas = span_nas.find_element(By.XPATH, 'span')
        # span_nas.click()
        # root_dogovor = root.find_element(By.XPATH,
        #                                  "//span[@class='select2-container select2-container--default select2-container--open']")
        # input_dogovor = root_dogovor.find_element(By.XPATH, ".//input[@class='select2-search__field']")
        # center = 'г.'+center
        # input_dogovor.send_keys(center)

        # if self._sel_wait_el(By.XPATH,
        #                      "//li[@class='select2-results__option select2-results__option--highlighted' and text()='" + center + "']",
        #                      sec=5):
        #     li_center = root.find_element(By.XPATH,
        #                                      "//li[@class='select2-results__option select2-results__option--highlighted' and text()='" + center + "']")
        # elif self._sel_wait_el(By.XPATH, "//li[@class='select2-results__option' and text()='" + center + "']", sec=5):
        #     li_center = root.find_element(By.XPATH,
        #                                      "//li[@class='select2-results__option' and text()='" + center + "']")
        # else:
        #     raise ValueError("Время ожидания истекло: Должность не найдена")
        # li_center.click()

        apply = root.find_element(By.XPATH, "//input[@value='Сохранить']")
        apply.click()

        winlog.info(f"count_dopka {len(count_dopka)}")
        # TODO Тут если count равно 0 то надо коунт + 1
        if len(count_dopka) == 0:
            winlog.info("launch part for first dopka")
            self._sel_wait_el(By.XPATH, f'//div[@class="table-responsive"]/table/tbody//tr[{len(count_dopka) + 1}]/td[7]/div/button')
            root.find_element(By.XPATH, f'//div[@class="table-responsive"]/table/tbody//tr[{len(count_dopka) + 1}]/td[7]/div/button').click()
            self._sel_wait_el(By.XPATH, f'//a[@class="dropdown-item sendDopBtn"]')
            root.find_element(By.XPATH, f'//a[@class="dropdown-item sendDopBtn"]').click()
        else:
            winlog.info("launch part for second dopka")
            self._sel_wait_el(By.XPATH,
                              f'//div[@class="table-responsive"]/table/tbody//tr[{len(count_dopka)}]/td[7]/div/button')
            root.find_element(By.XPATH,
                              f'//div[@class="table-responsive"]/table/tbody//tr[{len(count_dopka)}]/td[7]/div/button').click()
            self._sel_wait_el(By.XPATH, f'(//a[@class="dropdown-item sendDopBtn"])[{len(count_dopka)}]')
            root.find_element(By.XPATH, f'(//a[@class="dropdown-item sendDopBtn"])[{len(count_dopka)}]').click()
        self._sel_wait_el(By.XPATH, '//button[text()="OK"]')
        root.find_element(By.XPATH, '//button[text()="OK"]').click()
        self.ecp_priem(len(count_dopka))

        sleep(5)
        return 1, 0, 0

    def _sel_wait_el(self, by, selector, sec=60, appear=True):
        """Ожидание элемента появление или изчезновение подается через bool 'appear'"""
        driver = self.driver
        time.sleep(0.3)
        try:
            if appear:
                WebDriverWait(driver, sec).until(ec.presence_of_element_located((by, selector)))
            else:
                WebDriverWait(driver, sec).until_not(ec.presence_of_element_located((by, selector)))
            return True
        except:
            return False
        finally:
            time.sleep(0.2)

    def _search_iin_create1(self, root, iin, driver):
        self.anchor.click()

        input_iin = root.find_element(By.XPATH, ".//input[@name='iin']")
        input_iin.send_keys(iin)
        button_iin = root.find_element(By.XPATH,
                                       './/button[@type="submit" and text()="Найти"]')
        button_iin.click()

        dog = driver.find_element_by_class_name("item-list")
        dog_i = dog.find_element_by_xpath('div/div/a')
        link = dog_i.get_attribute('href')
        winlog.info(link)
        dog_i.click()

        dop_dr = self.driver
        dop_dr.get(link)
        dop = WebDriverWait(dop_dr, 3).until(ec.presence_of_element_located((By.LINK_TEXT, 'Добавить доп. соглашение')))

        count_exist_dop_sogl = self.driver.find_elements(By.XPATH, '//div[@class="table-responsive"]/table/tbody//tr')
        link1 = dop.get_attribute('href')
        winlog.info(link1)
        dop.click()
        return link1, count_exist_dop_sogl

    def close(self):
        self.driver.quit()

# TODO class ======================================================================================================================================


class EnbekDismissal:
    def __init__(self):
        self.robot_path = os.getcwd()
        self.driver = None
        self.downloads_path = os.path.join(os.environ['USERPROFILE'], "Downloads\\").replace("/", "\\")
        self.temp_path = self.robot_path + "Temp\\Enbek"
        self.path = {"downloads": {"title": "Downloads", "path": self.downloads_path, "dir": self.downloads_path[:-1]},
                     "temp": {"title": "Temp", "path": self.temp_path, "dir": self.temp_path[:-1]},
                     "enbek_files": {"title": "Enbek_files", "path": self.temp_path + "Enbek_files\\",
                                     "dir": self.temp_path + "Enbek_files"}}
        self.url = {
            "login": "https://www.enbek.kz/docs/ru/user",
            "list": "https://www.enbek.kz/ru/cabinet/dogovor/list/good",
            "add": "https://www.enbek.kz/ru/cabinet/dogovor/add",
        }
        # Data containers
        self.anchor = None
        print("\t")

    def ecp_dismissal(self):
        self._sel_wait_el(By.XPATH, '//button[text()="OK"]')
        self.driver.find_element(By.XPATH, '//button[text()="OK"]').click()

        file_name_input = bySelector([{"title": "Электронная биржа труда | ID ДОГОВОРА (E-HR): - Google Chrome", "backend": "uia"},
                                      {"depth_start": 3, "depth_end": 3, "title": "File name:",
                                       "control_type": "Edit"}])
        file_name_input.wait_appear()
        file_name_input.set_focus()
        sleep(0.5)
        file_name_input.set_edit_text(
            r"C:\Users\robot.drp\Desktop\Rasul\HCSBKKZ_robot\Tools\ecp\GOSTKNCA_42e6bd1a0979ff3747e9a35207bbcfbc4afee6f5.p12")
        sleep(0.5)

        open_button = bySelector([{"title": "Электронная биржа труда | ID ДОГОВОРА (E-HR): - Google Chrome", "backend": "uia"},
                                  {"depth_start": 2, "depth_end": 2, "title": "Open", "control_type": "Button"}])
        open_button.click()

        driver = self.driver
        self._sel_wait_el(By.XPATH, '(//input[@name="ecpPassword"])[1]')
        self._sel_wait_el(By.XPATH, '(//input[@name="ecpPassword"])[2]')
        driver.find_element(By.XPATH, '(//input[@name="ecpPassword"])[2]').send_keys("test_password")
        driver.find_element(By.XPATH, '(//button[@class="button-primary complaint-btn-submit form-sign-modal-send"])[2]').click()

    def _sel_init(self):
        """Запуск драйвера и логин в enbek. Объект драйвера создается именно здесь"""
        # Driver init
        try:
            self.driver = webdriver.Chrome(ChromeDriverManager().install())
        except Exception as e:
            winlog.warning(f"{datetime.datetime.today()} {'-' * 10} _sel_init_3 {e}")
            self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.switch_to.window(self.driver.current_window_handle)
        driver = self.driver
        driver.get(self.url["list"])
        login = "//input[@placeholder='Логин или E-mail']"
        WebDriverWait(driver, 60).until(ec.presence_of_element_located((By.XPATH, login)))
        driver.find_element(By.XPATH, login).send_keys("madieva.an@hcsbk.kz")
        passw = "//input[@placeholder='Пароль']"
        driver.find_element(By.XPATH, passw).send_keys("Ghjdthrf_2021")
        driver.find_element(By.XPATH, passw).send_keys(Keys.RETURN)
        driver.get(self.url["list"])
        return self.driver

    def terminate_dog(self, data):
        if not self.driver:
            self._sel_init()

        winlog.info('first one')
        driver = self._check_del_dog(data)
        winlog.info('second')
        self._data_append_del(data, driver)

    def _check_del_dog(self, data):
        driver = self.driver
        driver.get(self.url["list"])
        if self._sel_wait_el(By.XPATH, "//a[text()[contains(., 'Добавить')]]"):
            self.anchor = driver.find_element(By.XPATH, '//strong[text()="Договоры"]')
            root = driver.find_element(By.CSS_SELECTOR, ".content")

            fio = data[0]
            driver_common = self._search_iin_create2(root, fio, driver)
            winlog.info("self._check_dogovor > done")
            return driver_common
        else:
            winlog.info("self._check_dogovor > done")
            raise ValueError("Время ожидания истекло: https://www.enbek.kz/ru/cabinet/dogovor/list")

    def _data_append_del(self, data, root):
        date_rast_dog = WebDriverWait(root, 3).until(ec.presence_of_element_located((By.XPATH, "//input[@name='dateCutDogovor']")))
        date_rast_dog.click()
        winlog.info("1")
        root.execute_script('document.getElementsByName("dateCutDogovor")[0].removeAttribute("readonly")')
        date_rast_dog.send_keys(data[1])
        winlog.info("2")
        prichina = root.find_element(By.XPATH, "//div[@class='prich']/div/span/span/span")
        prichina.click()
        # prich = 'По соглашению сторон'
        prich = data[2]
        prich = prich[0].upper() + prich[1:]
        input_prichina = root.find_element(By.XPATH, ".//input[@class='select2-search__field']")
        input_prichina.send_keys(prich)
        winlog.info("3")
        print("prichina ", prich)
        if self._sel_wait_el(By.XPATH,
                             f'//li[@class="select2-results__option select2-results__option--highlighted" and text()="{prich.strip()} "]',
                             sec=5):
            li_prichina = root.find_element(By.XPATH, f'//li[@class="select2-results__option select2-results__option--highlighted" and text()="{prich.strip()} "]')
        elif self._sel_wait_el(By.XPATH,
                             f'//li[@class="select2-results__option select2-results__option--highlighted" and text()="{prich.strip()}"]',
                             sec=5):
            li_prichina = root.find_element(By.XPATH, f'//li[@class="select2-results__option select2-results__option--highlighted" and text()="{prich.strip()}"]')
        elif self._sel_wait_el(By.XPATH, "//li[@class='select2-results__option' and text()='" + prich.strip() + "']", sec=5):
            li_prichina = root.find_element(By.XPATH,
                                           "//li[@class='select2-results__option' and text()='" + prich.strip() + " ']")
        else:
            raise ValueError("Время ожидания истекло: Причина увольнения не совпадает")
        li_prichina.click()
        if not self._sel_wait_el(By.CSS_SELECTOR, "span.select2-container--open input.select2-search__field",
                                 appear=False):
            raise ValueError("Время ожидания истекло: Должность не выбрана")

        submit = root.find_element(By.XPATH, "//button[text()='Расторгнуть']")
        submit.click()

        self.ecp_dismissal()

        # while True:
        #     pass
        # data["cause"]

    def _sel_wait_el(self, by, selector, sec=60, appear=True):
        """Ожидание элемента появление или изчезновение подается через bool 'appear'"""
        driver = self.driver
        time.sleep(0.3)
        try:
            if appear:
                WebDriverWait(driver, sec).until(ec.presence_of_element_located((by, selector)))
            else:
                WebDriverWait(driver, sec).until_not(ec.presence_of_element_located((by, selector)))
            return True
        except:
            return False
        finally:
            time.sleep(0.2)

    def _search_iin_create2(self, root, fio, driver):
        self.anchor.click()
        winlog.info('do_iin')
        input_iin = root.find_element(By.XPATH, '//input[@name="fam"]')
        input_iin.send_keys(fio.split(" ")[0].strip())
        button_iin = root.find_element(By.XPATH,
                                       './/button[@type="submit" and text()="Найти"]')
        button_iin.click()
        status_loading_page = driver.execute_script("return document.readyState")
        print("status_loading_page", status_loading_page)
        while str(status_loading_page) != "complete":
            print("status_loading_page", status_loading_page)

        dog = driver.find_elements_by_xpath('//div[@class="item-list d-flex align-items-center"]')
        link_number = ""
        print("fio", fio)
        for i, el in enumerate(dog):
            print(i, "el.text", el.text)
            if fio.lower().strip() in el.text.lower():
                print("exist fio")
                start_link_number = 0
                end_link_number = 0
                flag_find_start_and_end = 0
                for i, word in enumerate(el.text):
                    # print("i", i, "word", word)
                    if word.isdigit():
                        start_link_number = i
                        for j, word_ in enumerate(el.text):
                            if j < i:
                                continue
                            if not word_.isdigit():
                                end_link_number = j
                                flag_find_start_and_end = 1
                                link_number = el.text[start_link_number:end_link_number].strip()
                                break
                    if flag_find_start_and_end:
                        break
        print("link_number", link_number)
        if not link_number:
            raise ValueError("Не найден действующий договор")
        winlog.info(f"link number : {link_number}")
        # tag_a = driver.find_element(By.XPATH, f'(//a[@href="/ru/cabinet/dogovor/{link_number}"])[1]')
        #
        # winlog.info(tag_a)
        # link = tag_a.get_attribute('href')
        # winlog.info(link)
        # tag_a.click()
        driver.get(f"https://enbek.kz/ru/cabinet/dogovor/{link_number}")
        winlog.info("2--")

        # dop_dr = self.driver
        # dop_dr.get(link)
        time.sleep(2)
        rast = WebDriverWait(driver, 3).until(ec.presence_of_element_located((By.XPATH, "//button[text()='Расторгнуть']")))
        # rast = dop_dr.find_element(By.XPATH, "//button[text()='Расторгнуть']")
        rast.click()
        return driver

    def close(self):
        self.driver.quit()


def enbek_priem():
    """Priem na rabotu"""

    # чтобы новые должности добавленные в эксель вручную, попадали сразу в json (оттуда потом берем наименование должности для енбека )
    extract_dolzhnosts_from_excel()



    dir_to_jsons = os.path.join(os.getcwd(), "jsons")

    all_jsons_list = list(pathlib.Path(dir_to_jsons).glob('*.json'))

    priem_json_list = []
    for file in all_jsons_list:
        if "priem" in str(file):
            priem_json_list.append(str(file))
    if priem_json_list:

        driver = Enbek()
        driver._sel_init()

        for file in priem_json_list:
            try:
                winlog.info(file)
                with open(file, encoding="utf-8") as file_:
                    data = json.load(file_)
                driver.create_dogovor(data)
                os.replace(file, os.path.join(os.getcwd(), "jsons", "done", file.split("\\")[-1]))
            except Exception as e:
                winlog.warning(f"{datetime.datetime.today()} enbek_priem {e}")
                os.replace(file, os.path.join(os.getcwd(), "jsons", "failed", file.split("\\")[-1]))
                # driver.create_dogovor(data)

        try:
            driver.driver.quit()
        except Exception as e:
            winlog.info(f"{datetime.datetime.today()} {'-' * 10} enbek_priem finally close driver {e}")
    else:
        winlog.info(f"enbek_priem do not launch. matching files do not found ")

    # driver.driver.execute_script("""Swal.fire({
    #   position: 'center',
    #   icon: 'success',
    #   title: 'Поздравляем. Процесс приема успешно подошел к концу',
    #   showConfirmButton: false,
    #   timer: 5000
    # })""")
    sleep(5)
    # input("press enter for close")


def enbek_perevod():
    # чтобы новые должности добавленные в эксель вручную, попадали сразу в json (оттуда потом берем наименование должности для енбека )
    EnbekPerevod.extract_data_from_excel_for_perevod()

    dir_to_jsons = os.path.join(os.getcwd(), "jsons")

    all_jsons_list = list(pathlib.Path(dir_to_jsons).glob('*.json'))

    perevod_json_list = []
    for file in all_jsons_list:
        if "perevod" in str(file):
            perevod_json_list.append(str(file))

    # with open((os.path.join(os.getcwd(), 'data_for_perevod.json'))) as colvir_json:
    #     code = json.load(colvir_json)
    #
    # data_perevod_dict = code['data_for_perevod'][0]

    not_found_list = []

    if perevod_json_list:
        driver = EnbekPerevod()
        driver._sel_init()

        for file in perevod_json_list:
            try:
                with open(file, encoding="utf-8") as file_:
                    data = json.load(file_)
                    data_perevod_dict = data['data_for_perevod'][0]
                for tabel_num in data_perevod_dict:
                    winlog.info(f"data_perevod {data}")
                    winlog.info(f"data_perevod_dict {data_perevod_dict}")
                    # not_found = []
                    status_dolzh_found, department, dolzhnost = driver.create_dop_sogl(data_perevod_dict[tabel_num])
                    # not_found.append(department)
                    # not_found.append(dolzhnost)
                    # not_found_list.append(not_found)
                    if not status_dolzh_found:
                        with open(file, encoding="utf-8") as file_:
                            data = json.load(file_)
                            print("this is data", data)
                            data["FAILED_REASON"] = f"NOT FOUND Department: {department} dolzhnost: {dolzhnost}"
                        with open(file, "w", encoding="utf-8") as file_:
                            json.dump(data, file_, ensure_ascii=False, indent=4)
                        winlog.info(f"{datetime.datetime.today()} {'-' * 10} dolzhnost not found")
                        os.replace(file, os.path.join(os.getcwd(), "jsons", "failed", file.split("\\")[-1]))
                    else:
                        with open(file, encoding="utf-8") as file_:
                            data = json.load(file_)
                            winlog.info(f"this is data {data}")
                            data["SUCCESS"] = f"{datetime.datetime.today()} Department: {department} dolzhnost: {dolzhnost}"
                        with open(file, "w", encoding="utf-8") as file_:
                            json.dump(data, file_, ensure_ascii=False, indent=4)
                        os.replace(file, os.path.join(os.getcwd(), "jsons", "done", file.split("\\")[-1]))
            except Exception as e:
                with open(file, encoding="utf-8") as file_:
                    data = json.load(file_)
                    print("this is data", data)
                    data["FAILED_REASON"] = f"{datetime.datetime.today()} {'-' * 10} {e}"
                with open(file, "w", encoding="utf-8") as file_:
                    json.dump(data, file_, ensure_ascii=False, indent=4)
                winlog.warning(f"{datetime.datetime.today()} Exception load *perevod.json {e}")
                os.replace(file, os.path.join(os.getcwd(), "jsons", "failed", file.split("\\")[-1]))
        driver.close()
    else:
        winlog.info(f"{datetime.datetime.today()} {'-' * 10} Driver not launch. Do not found matching files for 'perevod' in json files")


def enbek_dismissal():
    dir_to_jsons = os.path.join(os.getcwd(), "jsons")

    all_jsons_list = list(pathlib.Path(dir_to_jsons).glob('*.json'))

    perevod_json_list = []
    for file in all_jsons_list:
        if "dismissal" in str(file):
            perevod_json_list.append(str(file))

    if perevod_json_list:
        driver = EnbekDismissal()
        driver._sel_init()

        for i, file in enumerate(perevod_json_list):
            try:
                with open(file, encoding="utf-8") as file_:
                    data_ = json.load(file_)
                data = data_["data_for_dismissal"]
                ##################################
                driver.terminate_dog(data=data)

                data_["SUCCESS_DISMISSAL"] = f"{datetime.datetime.today()}"

                with open(file, "w", encoding="utf-8") as file_:
                    json.dump(data_, file_, ensure_ascii=False, indent=4)

                os.replace(file, os.path.join(os.getcwd(), "jsons", "done", file.split("\\")[-1]))
            except ValueError:
                winlog.warning(f"{datetime.datetime.today()} {'-' * 10} enbek_dismissal {ValueError}")
                with open(file, encoding="utf-8") as file_:
                    data = json.load(file_)
                data_ = data["data_for_dismissal"][2]
                data[
                    "FAILED_REASON_DISMISSAL"] = f"{datetime.datetime.today()} {'-' * 10} REASON: Не найден действующий договор EXCEPTION: {ValueError}"
                with open(file, "w", encoding="utf-8") as file_:
                    json.dump(data, file_, ensure_ascii=False, indent=4)

                os.replace(file, os.path.join(os.getcwd(), "jsons", "failed", file.split("\\")[-1]))
            except Exception as e:
                print("Launch Exception as e")
                winlog.warning(f"{datetime.datetime.today()} {'-' * 10} enbek_dismissal {e}")
                with open(file, encoding="utf-8") as file_:
                    data = json.load(file_)
                data_ = data["data_for_dismissal"][2]
                data["FAILED_REASON_DISMISSAL"] = f"{datetime.datetime.today()} {'-' * 10} REASON: {data_} EXCEPTION: {e}"
                with open(file, "w", encoding="utf-8") as file_:
                    json.dump(data, file_, ensure_ascii=False, indent=4)

                os.replace(file, os.path.join(os.getcwd(), "jsons", "failed", file.split("\\")[-1]))

        driver.close()
    else:
        winlog.info(f"NOT FOUND JSON FILES FOR LAUNCH DISMISSAL PROCESS")
    #
    # with open((os.path.join(os.getcwd(), 'data_for_dismissal.json'))) as colvir_json:
    #     code = json.load(colvir_json)

    # Для теста ( других актуальных данных не было, полученный из эксельки с колвира список людей были уже уволенными )
    # TODO Надо добавить чек по ФИО ( т.к. ИИН нету и поиск по фамилии, а фамилии естесственно могут совпадать )
    # code = [["ПЛОТНИКОВА СВЕТЛАНА ВЛАДИМИРОВНА", "15.10.2021", "по соглашению сторон"]]
    # for data in code:
    #     driver.terminate_dog(data=data)


